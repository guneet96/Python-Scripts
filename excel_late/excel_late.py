import openpyxl
import pandas as pd
w = openpyxl.load_workbook('data.xlsx')
sheet = w.get_sheet_by_name('Sheet1')


#column
i = sheet['B']

d = dict()
d = {"india":1, "pakistan":2, "canada":3, "brazil":4, "soviet":5}

for k in range(1,sheet.max_row + 1):
	name_country = sheet.cell(row=k, column = 2).value;
	print sheet.cell(row=k, column = 2).value
	if name_country in d:
		sheet.cell(row=k, column = 3).value = d.get(name_country)
		print d.get(name_country)

w.save('answer.xlsx')	
